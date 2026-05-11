"""聊天和任务管理路由"""
from flask import Blueprint, request, jsonify, send_file, Response, stream_with_context
import json
import os
import shlex
import sys
import threading
import subprocess
import re as _re
import time as _time
import uuid as _uuid
from datetime import datetime
from pathlib import Path
from werkzeug.utils import secure_filename
from state import (
    tasks, _session_docs, _science_agent_jobs, _lit_jobs, _chat_jobs,
    UPLOAD_FOLDER_CHAT, SCIENCE_WORKSPACE_ROOT, _PROJECT_ROOT,
)
from helpers import (
    USER_PROFILE_MD,
    append_user_profile_to_system,
    get_user_profile_context,
    get_llm_config,
    llm_call,
    llm_stream,
    inject_workspace_context,
    get_workspace_config,
    get_kb_instance,
    safe_child_path,
)

bp = Blueprint('chat', __name__)

_RAG_COMPLETE_ONLY_RULE = (
    "===== RAG 引用规则 / RAG citation rule =====\n"
    "下面的上传文档或知识库内容可能来自切片。回答时只能引用或复述完整句子、完整段落、完整列表项、完整表格行和完整代码块；"
    "疑似截断的半句话、半个命令、未闭合代码块只能作为检索线索，不要当作事实或原文输出。\n"
)


def _drop_incomplete_fenced_tail(text: str) -> str:
    lines = str(text or "").splitlines()
    in_fence = False
    fence_char = ""
    fence_len = 0
    open_idx = -1
    for i, line in enumerate(lines):
        m = _re.match(r"^\s*([`~]{3,})", line)
        if not m:
            continue
        marker = m.group(1)
        char = marker[0]
        if any(ch != char for ch in marker):
            continue
        if not in_fence:
            in_fence = True
            fence_char = char
            fence_len = len(marker)
            open_idx = i
        elif char == fence_char and len(marker) >= fence_len:
            in_fence = False
            fence_char = ""
            fence_len = 0
            open_idx = -1
    if in_fence and open_idx >= 0:
        return "\n".join(lines[:open_idx]).rstrip()
    return str(text or "").rstrip()


# ── Persistent chat history ────────────────────────────────────────────────

CHAT_HISTORY_DIR = _PROJECT_ROOT / "seismo_rag" / "chat_history"
CHAT_HISTORY_JSON = CHAT_HISTORY_DIR / "conversations.json"
USER_PROFILE_ARCHIVE_DIR = _PROJECT_ROOT / "seismo_rag" / "user_profiles"
USER_PROFILE_SOURCE_JSON = _PROJECT_ROOT / "seismo_rag" / "user_profile_source.json"
CHAT_HISTORY_DIR.mkdir(parents=True, exist_ok=True)
USER_PROFILE_ARCHIVE_DIR.mkdir(parents=True, exist_ok=True)


def _clean_conversation_id(value: str) -> str:
    return "".join(ch for ch in str(value or "default") if ch.isalnum() or ch in "_-")[:80] or "default"


def _safe_chat_pdf_upload_path(filename: str | None, session_id: str, upload_id: str) -> tuple[Path, str]:
    raw_name = Path(filename or "").name
    if Path(raw_name).suffix.lower() != ".pdf":
        raise ValueError("Only PDF files are supported")
    safe_stem = secure_filename(Path(raw_name).stem) or "upload"
    safe_name = f"{safe_stem}.pdf"
    tmp_name = f"{_clean_conversation_id(session_id)}_{upload_id}_{safe_name}"
    return safe_child_path(UPLOAD_FOLDER_CHAT, tmp_name), safe_name


def _safe_pdf_upload_path(filename: str | None, session_id: str) -> tuple[Path, str]:
    """Compatibility wrapper used by tests and older upload code paths."""
    upload_id = _uuid.uuid4().hex[:10]
    return _safe_chat_pdf_upload_path(filename, session_id, upload_id)


def _extract_session_pdf_chunks(path: str | Path, upload_id: str, doc_name: str) -> tuple[list, list]:
    """Extract temporary chat-PDF text chunks using the current RAG extractors."""
    sys.path.insert(0, str(Path(__file__).parent))
    from rag_extractors import extract_text, chunk_text

    pages = extract_text(str(path))
    chunks = []
    for page_idx, page_text in pages:
        for c in chunk_text(page_text, chunk_size=600):
            chunks.append({
                "page": page_idx + 1,
                "text": c,
                "doc_name": doc_name,
                "upload_id": upload_id,
            })
    return pages, chunks


def _science_read_file_preview(path: Path, max_chars: int = 4000) -> str:
    """Best-effort text/table preview for scientific-analysis project files."""
    ext = path.suffix.lower()
    try:
        if ext in {".png", ".jpg", ".jpeg", ".svg", ".tif", ".tiff"}:
            try:
                if ext == ".svg":
                    data = path.read_text(encoding="utf-8", errors="ignore")[:max_chars]
                    return (
                        "[image/svg evidence candidate]\n"
                        f"path: {path}\n"
                        f"size_bytes: {path.stat().st_size}\n"
                        "Use a vision-capable model or image parser if quantitative visual evidence is needed.\n"
                        f"svg_preview:\n{data}"
                    )[:max_chars].strip()
                from PIL import Image  # type: ignore
                with Image.open(str(path)) as im:
                    return (
                        "[image evidence candidate]\n"
                        f"path: {path}\n"
                        f"format: {im.format}\n"
                        f"mode: {im.mode}\n"
                        f"width: {im.width}\n"
                        f"height: {im.height}\n"
                        f"size_bytes: {path.stat().st_size}\n"
                        "Use a vision-capable model for visual interpretation; otherwise treat this as an image artifact."
                    )
            except Exception as exc:
                return f"[image metadata unavailable: {exc}]"
        if ext in {".pdf", ".docx", ".md", ".txt", ".text", ".rst", ".html", ".htm"}:
            sys.path.insert(0, str(Path(__file__).parent))
            from rag_extractors import extract_text
            pages = extract_text(str(path))
            text = "\n\n".join(str(t or "") for _, t in pages)
            return text[:max_chars].strip()
        if ext == ".doc":
            try:
                proc = subprocess.run(
                    ["textutil", "-convert", "txt", "-stdout", str(path)],
                    capture_output=True,
                    text=True,
                    timeout=15,
                    check=False,
                )
                if proc.stdout.strip():
                    return proc.stdout[:max_chars].strip()
            except Exception:
                pass
            data = path.read_bytes()[:max_chars * 2]
            return data.decode("utf-8", errors="ignore")[:max_chars].strip()
        if ext in {".csv", ".tsv", ".dat", ".json", ".yaml", ".yml", ".bib", ".tex", ".py", ".sh"}:
            data = path.read_bytes()[:max_chars * 2]
            return data.decode("utf-8", errors="ignore")[:max_chars].strip()
        if ext in {".xlsx", ".xls"}:
            try:
                from openpyxl import load_workbook  # type: ignore
                wb = load_workbook(str(path), read_only=True, data_only=True)
                parts = []
                for ws in wb.worksheets[:3]:
                    parts.append(f"[sheet] {ws.title}")
                    for row in ws.iter_rows(max_row=12, values_only=True):
                        parts.append("\t".join("" if v is None else str(v) for v in row[:12]))
                return "\n".join(parts)[:max_chars].strip()
            except Exception as exc:
                return f"[xlsx preview unavailable: {exc}]"
        if ext in {".h5", ".hdf5"}:
            try:
                import h5py  # type: ignore
                lines = []
                with h5py.File(str(path), "r") as h5:
                    def visit(name, obj):
                        shape = getattr(obj, "shape", "")
                        dtype = getattr(obj, "dtype", "")
                        lines.append(f"{name} shape={shape} dtype={dtype}")
                    h5.visititems(visit)
                return "\n".join(lines[:80])[:max_chars].strip()
            except Exception as exc:
                return f"[hdf5 metadata unavailable: {exc}]"
        if ext == ".nc":
            try:
                from netCDF4 import Dataset  # type: ignore
                with Dataset(str(path)) as ds:
                    lines = ["dimensions: " + ", ".join(f"{k}={len(v)}" for k, v in ds.dimensions.items())]
                    lines.append("variables: " + ", ".join(list(ds.variables.keys())[:80]))
                return "\n".join(lines)[:max_chars].strip()
            except Exception as exc:
                return f"[netcdf metadata unavailable: {exc}]"
    except Exception as exc:
        return f"[preview extraction failed: {exc}]"
    return ""


def _science_numeric_text_score(text: str) -> float:
    lines = [ln.strip() for ln in (text or "").splitlines() if ln.strip()][:60]
    if not lines:
        return 0.0
    numeric = 0
    for ln in lines:
        tokens = _re.split(r"[\s,;]+", ln)
        vals = 0
        for tok in tokens:
            try:
                float(tok)
                vals += 1
            except Exception:
                pass
        if vals >= 3:
            numeric += 1
    return numeric / max(len(lines), 1)


def _science_guess_file_role(path: Path, preview: str) -> str:
    name = path.name.lower()
    ext = path.suffix.lower()
    norm_parts = {str(part).lower() for part in path.parts}
    if (
        "parameter_optimization" in norm_parts
        or "parameter_optimization" in name
        or name in {"best_parameters.json", "optimization_history.csv", "optimization_report.md", "optimization_job_summary.json"}
        or "optimization" in name
    ):
        return "parameter_optimization_evidence"
    if "article" in path.parts and ext in {".pdf", ".aux", ".log", ".blg", ".bbl"}:
        return "article_template_output"
    if "article" in path.parts and ext in {".tex", ".bib", ".cls", ".sty"}:
        return "article_template_file"
    if any(k in path.name for k in ("数据说明", "说明", "字段", "README", "readme")) or "description" in name:
        return "data_description_or_project_notes"
    if ext == ".pdf":
        return "reference_paper_or_report"
    if ext in {".csv", ".tsv", ".xlsx", ".xls", ".parquet", ".h5", ".hdf5", ".nc"}:
        return "structured_data"
    if ext in {".sac", ".mseed"}:
        return "waveform_data"
    if ext in {".py", ".ipynb", ".sh"}:
        return "code_or_processing_script"
    if ext in {".png", ".jpg", ".jpeg", ".svg"}:
        return "figure_or_image_evidence"
    if ext in {".txt", ".dat"} and _science_numeric_text_score(preview) > 0.35:
        return "numeric_text_data"
    if ext in {".md", ".txt", ".doc", ".docx", ".rst", ".html", ".htm"}:
        return "document_or_notes"
    return "unknown_project_file"


def _science_build_file_profiles(root: Path, max_files: int = 160) -> list[dict]:
    """Scan project files and attach previews/role hints for the science agent prompt."""
    if not root.exists():
        return []
    skip_dirs = {".git", "__pycache__", ".pytest_cache", "node_modules"}
    candidates = []
    for p in root.rglob("*"):
        if not p.is_file():
            continue
        if any(part in skip_dirs for part in p.parts):
            continue
        if "outputs" in p.parts and "science_analysis_agent" in p.parts:
            continue
        if p.suffix.lower() not in _SCI_ALLOWED_EXTS:
            continue
        candidates.append(p)

    def priority(p: Path):
        ext = p.suffix.lower()
        name = p.name.lower()
        score = 5
        if any(k in p.name for k in ("数据说明", "说明", "字段")) or "readme" in name:
            score = 0
        elif "science_analysis_inputs" in p.parts or "parameter_optimization" in str(p).lower() or "optimization" in name:
            score = 1
        elif ext in {".csv", ".tsv", ".xlsx", ".xls", ".txt", ".dat", ".json", ".h5", ".hdf5", ".nc", ".sac", ".mseed"}:
            score = 1
        elif ext == ".pdf":
            score = 2
        elif ext in {".doc", ".docx", ".md", ".rst"}:
            score = 3
        elif ext in {".py", ".ipynb", ".sh"}:
            score = 4
        return (score, str(p).lower())

    profiles = []
    for p in sorted(candidates, key=priority)[:max_files]:
        preview = _science_read_file_preview(p)
        try:
            rel = str(p.relative_to(root))
        except Exception:
            rel = str(p)
        profiles.append({
            "path": rel,
            "abs_path": str(p),
            "suffix": p.suffix.lower() or "file",
            "size": p.stat().st_size,
            "role_hint": _science_guess_file_role(p, preview),
            "preview": preview[:2200],
            "numeric_score": round(_science_numeric_text_score(preview), 2),
        })
    return profiles


def _science_profiles_to_prompt(profiles: list[dict], role_summary: str = "") -> str:
    if not profiles:
        return "No readable project files were found."
    lines = []
    if role_summary:
        lines += ["===== LLM file-role assessment =====", role_summary.strip(), ""]
    lines.append("===== Project file profiles with previews =====")
    for idx, item in enumerate(profiles, 1):
        preview = (item.get("preview") or "").strip()
        preview = preview if preview else "[no text preview extracted; use filename, suffix, metadata, or multimodal tools if needed]"
        lines += [
            f"\n### [{idx}] {item.get('path')}",
            f"- suffix: {item.get('suffix')}  size: {item.get('size')} bytes",
            f"- rough_role_hint: {item.get('role_hint')}  numeric_score: {item.get('numeric_score')}",
            "- preview:",
            preview[:2200],
        ]
    return "\n".join(lines)


def _science_llm_role_summary(profiles: list[dict], data_description: str = "") -> str:
    """Ask the configured LLM to classify project files. Returns raw text, no JSON dependency."""
    if not profiles:
        return ""
    brief_parts = []
    for item in profiles[:60]:
        brief_parts.append(
            f"文件: {item.get('path')}\n"
            f"后缀: {item.get('suffix')}  初步提示: {item.get('role_hint')}\n"
            f"内容片段:\n{(item.get('preview') or '')[:900]}"
        )
    messages = [
        {"role": "system", "content": (
            "你是科研数据管家。请根据文件名、扩展名和内容片段判断每个文件在科研项目中的作用。"
            "不要编造未出现的信息；不确定时写“待验证”。直接输出 Markdown 表格和简短建议，不要输出 JSON。"
        )},
        {"role": "user", "content": (
            "用户提供的数据说明：\n"
            f"{data_description or '(未提供)'}\n\n"
            "请把下面文件分为：数据、数据说明/字段说明、参考论文、代码/脚本、图像/表格、其他。"
            "同时指出后续分析应优先读取哪些文件。\n\n"
            + "\n\n---\n\n".join(brief_parts)
        )},
    ]
    try:
        return llm_call(messages, get_llm_config(), max_tokens=1800)
    except Exception as exc:
        return f"[LLM file-role assessment unavailable: {exc}]"


def _science_resolve_workspace_root(value: str | Path | None, session_id: str = "default_science") -> Path:
    """Resolve Science Agent workspaces consistently, with legacy upload fallback."""
    raw = str(value or "").strip()
    if not raw:
        return SCIENCE_WORKSPACE_ROOT / _clean_conversation_id(session_id)
    p = Path(raw).expanduser()
    if p.is_absolute():
        return p
    repo_candidate = (_PROJECT_ROOT / p).resolve()
    legacy_candidate = (Path(__file__).parent.parent / p).resolve()
    if legacy_candidate.exists():
        repo_has_files = repo_candidate.exists() and any(x.is_file() for x in repo_candidate.rglob("*"))
        legacy_has_files = any(x.is_file() for x in legacy_candidate.rglob("*"))
        if not repo_candidate.exists() or (legacy_has_files and not repo_has_files):
            return legacy_candidate
    return repo_candidate


def _science_normalize_markdown_text(text: str) -> str:
    """Normalize common LaTeX fragments before web/PDF rendering."""
    value = str(text or "")
    value = _re.sub(r"M\$_\\(?:text|mathrm)\{([A-Za-z]+)\}\$", r"$M_{\\mathrm{\1}}$", value)
    value = _re.sub(r"M\$_\{([A-Za-z]+)\}\$", r"$M_{\1}$", value)
    value = _re.sub(r"M\$_([A-Za-z]+)\$", r"$M_{\1}$", value)
    return _drop_incomplete_fenced_tail(value)


def _science_rewrite_markdown_image_paths(markdown_text: str, output_dir: Path, figures: list | None = None) -> str:
    """Resolve Markdown image paths to local files so frontend and PDF export can load them."""
    output_dir = Path(output_dir).expanduser()
    by_name: dict[str, Path] = {}
    for item in figures or []:
        try:
            p = Path(str(item)).expanduser()
            if p.exists() and p.is_file():
                by_name[p.name] = p.resolve()
        except Exception:
            continue
    try:
        for p in output_dir.rglob("*"):
            if p.suffix.lower() in {".png", ".jpg", ".jpeg", ".svg"} and p.is_file():
                by_name.setdefault(p.name, p.resolve())
    except Exception:
        pass

    def _replace(match: _re.Match) -> str:
        alt, raw_href = match.group(1), match.group(2).strip()
        if not raw_href or raw_href.startswith(("http://", "https://", "data:")):
            return match.group(0)
        href = raw_href.strip("<>")
        candidate = Path(href).expanduser()
        if candidate.is_absolute() and candidate.exists():
            resolved = candidate.resolve()
        else:
            resolved = by_name.get(Path(href).name)
            if resolved is None:
                local = output_dir / href
                if local.exists():
                    resolved = local.resolve()
        if not resolved:
            return match.group(0)
        title = ""
        if " " in raw_href and raw_href.count('"') >= 2:
            title = " " + raw_href[raw_href.find('"'):]
        return f"![{alt}]({resolved}{title})"

    text = _re.sub(r"!\[([^\]]*)\]\(([^)\n]+)\)", _replace, markdown_text)

    inserted_names = {
        Path(m.group(1).strip().strip("<>")).name
        for m in _re.finditer(r"!\[[^\]]*\]\(([^)\n]+)\)", text)
    }
    lines = text.splitlines()
    for name, resolved in sorted(by_name.items()):
        if name in inserted_names or name not in text:
            continue
        image_line = f"![{Path(name).stem}]({resolved})"
        for idx, line in enumerate(lines):
            if name in line:
                lines[idx + 1:idx + 1] = ["", image_line, ""]
                inserted_names.add(name)
                break
    return "\n".join(lines)


def _science_finalize_result_artifacts(result: dict, output_dir: Path, progress_cb=None) -> None:
    """Create render-ready Markdown/HTML/PDF artifacts for Science Agent outputs."""
    if not isinstance(result, dict):
        return
    output_dir = Path(output_dir).expanduser()
    output_dir.mkdir(parents=True, exist_ok=True)
    md_text = result.get("markdown_paper") or ""
    md_path = Path(str(result.get("markdown_paper_path") or "")).expanduser()
    if not md_text and md_path.exists():
        try:
            md_text = md_path.read_text(encoding="utf-8")
        except Exception:
            md_text = ""
    if not md_text:
        return

    md_text = _science_normalize_markdown_text(md_text)
    md_text = _science_rewrite_markdown_image_paths(md_text, output_dir, result.get("generated_figures") or [])
    result["markdown_paper"] = md_text

    rendered_md = output_dir / "science_paper_rendered.md"
    try:
        rendered_md.write_text(md_text, encoding="utf-8")
        result["markdown_paper_rendered_path"] = str(rendered_md)
    except Exception as exc:
        result["markdown_render_error"] = str(exc)

    try:
        import markdown as _markdown
        body = _markdown.markdown(
            md_text,
            extensions=["extra", "tables", "fenced_code", "sane_lists"],
            output_format="html5",
        )
        html_path = output_dir / "science_paper_rendered.html"
        html_path.write_text(
            "<!doctype html><html><head><meta charset='utf-8'>"
            "<style>body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;line-height:1.65;"
            "max-width:920px;margin:32px auto;padding:0 28px;color:#222}h1{font-size:28px;line-height:1.25}"
            "h2{font-size:22px;margin-top:1.4em}h3{font-size:18px;margin-top:1.2em}"
            "img{max-width:100%;height:auto;display:block;margin:16px auto;border:1px solid #ddd}"
            "table{border-collapse:collapse;width:100%;margin:14px 0}th,td{border:1px solid #ddd;padding:6px 8px}"
            "pre{background:#f6f8fa;padding:12px;border-radius:8px;overflow:auto}code{background:#f3f4f6;padding:1px 4px}</style>"
            "</head><body>" + body + "</body></html>",
            encoding="utf-8",
        )
        result["paper_html"] = str(html_path)
    except Exception as exc:
        result["paper_html_error"] = str(exc)

    existing_pdf = str(result.get("latex_pdf") or result.get("paper_pdf") or "").strip()
    if existing_pdf and Path(existing_pdf).expanduser().exists():
        result["paper_pdf"] = existing_pdf
        return

    pdf_path = output_dir / "science_paper_rendered.pdf"
    rendered_md_path = result.get("markdown_paper_rendered_path")
    if not rendered_md_path:
        return
    pandoc = None
    try:
        import shutil as _shutil
        pandoc = _shutil.which("pandoc")
    except Exception:
        pandoc = None
    if not pandoc:
        result["paper_pdf_error"] = "pandoc not found; HTML export is available instead."
        return
    cmd = [
        pandoc,
        rendered_md_path,
        "-o",
        str(pdf_path),
        "--pdf-engine=xelatex",
        "-V",
        "CJKmainfont=Songti SC",
        "-V",
        "mainfont=Times New Roman",
    ]
    try:
        proc = subprocess.run(cmd, cwd=str(output_dir), capture_output=True, text=True, timeout=90, check=False)
        if proc.returncode != 0:
            cmd2 = [pandoc, rendered_md_path, "-o", str(pdf_path), "--pdf-engine=xelatex"]
            proc = subprocess.run(cmd2, cwd=str(output_dir), capture_output=True, text=True, timeout=90, check=False)
        if proc.returncode == 0 and pdf_path.exists() and pdf_path.stat().st_size > 0:
            result["paper_pdf"] = str(pdf_path)
            if progress_cb:
                progress_cb({"phase": "paper", "message": f"已生成 PDF: {pdf_path}"})
        else:
            result["paper_pdf_error"] = (proc.stderr or proc.stdout or "pandoc failed")[-1200:]
    except Exception as exc:
        result["paper_pdf_error"] = str(exc)


def _clean_conversation_title(value: str) -> str:
    title = str(value or "").strip()
    if title in {"新对话", "旧对话", "New chat", "Old chat"}:
        return ""
    return title[:120]


def _conversation_to_markdown(conv: dict) -> str:
    title = str(conv.get("title") or conv.get("id") or "Conversation")
    lines = [
        f"# {title}",
        "",
        f"- id: `{conv.get('id', '')}`",
        f"- created_at: {conv.get('createdAt', '')}",
        f"- updated_at: {conv.get('updatedAt', '')}",
        "",
    ]
    for m in conv.get("messages") or []:
        role = str(m.get("role") or "assistant")
        content = str(m.get("content") or "").strip()
        if not content:
            continue
        lines.extend([f"## {role}", "", content, ""])
    return "\n".join(lines).strip() + "\n"


def _load_persistent_conversations() -> dict:
    if not CHAT_HISTORY_JSON.exists():
        return {"conversations": [], "active_id": "", "projects": [], "active_project_id": ""}
    try:
        data = json.loads(CHAT_HISTORY_JSON.read_text(encoding="utf-8"))
        if not isinstance(data, dict):
            return {"conversations": [], "active_id": ""}
        convs = data.get("conversations") or []
        if not isinstance(convs, list):
            convs = []
        projects = data.get("projects") or []
        if not isinstance(projects, list):
            projects = []
        return {
            "conversations": convs,
            "active_id": data.get("active_id") or "",
            "projects": projects,
            "active_project_id": data.get("active_project_id") or "",
        }
    except Exception:
        return {"conversations": [], "active_id": "", "projects": [], "active_project_id": ""}


def _save_persistent_conversations(
    conversations: list,
    active_id: str = "",
    projects: list | None = None,
    active_project_id: str = "",
) -> None:
    CHAT_HISTORY_DIR.mkdir(parents=True, exist_ok=True)
    safe_convs = []
    for c in conversations[:200]:
        if not isinstance(c, dict):
            continue
        cid = _clean_conversation_id(c.get("id") or "")
        if not cid:
            continue
        safe = {
            "id": cid,
            "title": _clean_conversation_title(c.get("title") or ""),
            "project_id": _clean_conversation_id(c.get("project_id") or "") if c.get("project_id") else "",
            "createdAt": c.get("createdAt"),
            "updatedAt": c.get("updatedAt"),
            "history": (c.get("history") or [])[-80:],
            "messages": (c.get("messages") or [])[-200:],
            "docs": c.get("docs") or [],
        }
        safe_convs.append(safe)

    safe_projects = []
    for p in (projects or []):
        if not isinstance(p, dict):
            continue
        pid = _clean_conversation_id(p.get("id") or "")
        if not pid:
            continue
        safe_projects.append({
            "id": pid,
            "title": str(p.get("title") or "")[:120],
            "preface": str(p.get("preface") or ""),
            "prompt": str(p.get("prompt") or ""),
            "createdAt": p.get("createdAt"),
            "updatedAt": p.get("updatedAt"),
            "docs": p.get("docs") or [],
        })

    payload = {
        "active_id": _clean_conversation_id(active_id),
        "active_project_id": _clean_conversation_id(active_project_id or "") if active_project_id else "",
        "projects": safe_projects,
        "conversations": safe_convs,
    }
    tmp = CHAT_HISTORY_JSON.with_suffix(".json.tmp")
    tmp.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    tmp.replace(CHAT_HISTORY_JSON)

    for conv in safe_convs:
        md_path = CHAT_HISTORY_DIR / f"{conv['id']}.md"
        md_path.write_text(_conversation_to_markdown(conv), encoding="utf-8")


def _atomic_write_text(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(content, encoding="utf-8")
    tmp.replace(path)


def _delete_kb_proj_folder(kb, proj_folder: str) -> int:
    """Remove all KB docs in a logical folder before replacing it."""
    if not kb or not proj_folder:
        return 0
    removed = 0
    for doc in list(kb.list_docs()):
        if (getattr(doc, "proj_folder", "") or "") == proj_folder:
            try:
                if kb.delete_doc(doc.doc_id):
                    removed += 1
            except Exception:
                pass
    return removed


def _save_user_profile(content: str, conversations: list | None = None) -> None:
    """Persist the canonical profile plus a timestamped archive and source manifest."""
    now = datetime.now().isoformat(timespec="seconds")
    if not content.lstrip().startswith("#"):
        content = "# SAGE User Profile\n\n" + content.strip() + "\n"
    if "Updated:" not in content[:500]:
        content = content.rstrip() + f"\n\n---\nUpdated: {now}\n"

    _atomic_write_text(USER_PROFILE_MD, content)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    _atomic_write_text(USER_PROFILE_ARCHIVE_DIR / f"user_profile_{stamp}.md", content)
    if conversations is not None:
        source = {
            "updated_at": now,
            "profile_path": str(USER_PROFILE_MD),
            "archive_dir": str(USER_PROFILE_ARCHIVE_DIR),
            "n_conversations": len(conversations),
            "conversation_ids": [c.get("id") for c in conversations if isinstance(c, dict)],
        }
        tmp = USER_PROFILE_SOURCE_JSON.with_suffix(".json.tmp")
        tmp.write_text(json.dumps(source, ensure_ascii=False, indent=2), encoding="utf-8")
        tmp.replace(USER_PROFILE_SOURCE_JSON)


@bp.route('/api/chat/conversations', methods=['GET'])
def chat_conversations_get():
    data = _load_persistent_conversations()
    return jsonify({"ok": True, **data})


@bp.route('/api/chat/conversations', methods=['POST'])
def chat_conversations_save():
    data = request.json or {}
    conversations = data.get("conversations") or []
    if not isinstance(conversations, list):
        return jsonify({"ok": False, "error": "conversations must be a list"}), 400
    try:
        _save_persistent_conversations(
            conversations,
            data.get("active_id") or "",
            data.get("projects") or [],
            data.get("active_project_id") or "",
        )
        return jsonify({"ok": True, "path": str(CHAT_HISTORY_JSON)})
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 500


@bp.route('/api/chat/conversations/<conversation_id>', methods=['DELETE'])
def chat_conversation_delete(conversation_id):
    """Delete one persisted conversation and its Markdown mirror."""
    cid = _clean_conversation_id(conversation_id)
    data = _load_persistent_conversations()
    conversations = [c for c in data.get("conversations", []) if _clean_conversation_id(c.get("id")) != cid]
    active_id = data.get("active_id") or ""
    if active_id == cid:
        active_id = conversations[0].get("id", "") if conversations else ""
    try:
        _save_persistent_conversations(conversations, active_id, data.get("projects") or [], data.get("active_project_id") or "")
        md_path = CHAT_HISTORY_DIR / f"{cid}.md"
        if md_path.exists():
            md_path.unlink()
        _session_docs.pop(cid, None)
        _chat_jobs.pop(cid, None)
        return jsonify({"ok": True, "deleted": cid})
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 500


@bp.route('/api/chat/project/promote', methods=['POST'])
def chat_project_promote():
    """Persist a project's shared context and conversation summaries into the KB."""
    data = request.json or {}
    project = data.get("project") or {}
    conversations = data.get("conversations") or []
    pid = _clean_conversation_id(project.get("id") or "project")
    title = str(project.get("title") or pid)
    lines = [
        f"# Project: {title}",
        "",
        f"- Project ID: `{pid}`",
        f"- Source: SAGE Chat Project",
        "",
    ]
    if project.get("preface"):
        lines.extend(["## Preface / Research Background", "", str(project.get("preface")), ""])
    if project.get("prompt"):
        lines.extend(["## Shared Prompt", "", str(project.get("prompt")), ""])
    docs = project.get("docs") or []
    if docs:
        lines.extend(["## Project Literature", ""])
        for d in docs:
            lines.append(f"- {d.get('name')} ({d.get('n_chunks', 0)} chunks)")
        lines.append("")
    if conversations:
        lines.extend(["## Conversation Summaries", ""])
        for c in conversations[:50]:
            lines.append(f"### {c.get('title') or c.get('id')}")
            for m in (c.get("messages") or [])[-20:]:
                content = str(m.get("content") or "").strip()
                if content:
                    lines.append(f"- {m.get('role', '')}: {content[:1500]}")
            lines.append("")

    project_dir = CHAT_HISTORY_DIR / "projects"
    project_dir.mkdir(parents=True, exist_ok=True)
    md_path = project_dir / f"{pid}.md"
    md_path.write_text("\n".join(lines), encoding="utf-8")

    try:
        kb = get_kb_instance()
        if not kb:
            return jsonify({"ok": False, "error": "Knowledge base unavailable"}), 500
        logs = []
        proj_folder = f"chat_project/{pid}"
        removed_old = _delete_kb_proj_folder(kb, proj_folder)
        meta = kb.add_document(str(md_path), progress_cb=lambda m: logs.append(m), proj_folder=proj_folder, source_type="chat_project")
        return jsonify({
            "ok": True,
            "doc_id": meta.doc_id,
            "doc_name": meta.doc_name,
            "n_chunks": meta.n_chunks,
            "removed_old": removed_old,
            "path": str(md_path),
            "logs": logs[-20:],
        })
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 500


@bp.route('/api/chat/conversation/promote', methods=['POST'])
def chat_conversation_promote():
    """Persist the active chat conversation as one removable KB document."""
    data = request.json or {}
    conv = data.get("conversation") or {}
    cid = _clean_conversation_id(conv.get("id") or "conversation")
    title = _clean_conversation_title(conv.get("title") or "") or cid
    md = _conversation_to_markdown({**conv, "title": title})

    chat_dir = CHAT_HISTORY_DIR / "knowledge_chats"
    chat_dir.mkdir(parents=True, exist_ok=True)
    md_path = chat_dir / f"{cid}.md"
    md_path.write_text(md, encoding="utf-8")

    try:
        kb = get_kb_instance()
        if not kb:
            return jsonify({"ok": False, "error": "Knowledge base unavailable"}), 500
        logs = []
        proj_folder = f"chat/{cid}"
        removed_old = _delete_kb_proj_folder(kb, proj_folder)
        meta = kb.add_document(
            str(md_path),
            progress_cb=lambda m: logs.append(m),
            proj_folder=proj_folder,
            source_type="chat_conversation",
        )
        return jsonify({
            "ok": True,
            "doc_id": meta.doc_id,
            "doc_name": meta.doc_name,
            "n_chunks": meta.n_chunks,
            "removed_old": removed_old,
            "path": str(md_path),
            "logs": logs[-20:],
        })
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 500



# ── Task management ──────────────────────────────────────────────────────

_DEVICE_RE = _re.compile(r"^(cpu|mps|cuda(?::\d+)?)$")


def _require_text(data: dict, key: str, label: str) -> str:
    value = data.get(key)
    if value is None or not str(value).strip():
        raise ValueError(f"{label} required")
    return str(value).strip()


def _optional_text(data: dict, key: str, default: str) -> str:
    value = str(data.get(key, default) or "").strip()
    return value or default


def _validated_device(data: dict) -> str:
    device = _optional_text(data, "device", "cpu")
    if not _DEVICE_RE.match(device):
        raise ValueError("Invalid device; expected cpu, mps, cuda, or cuda:N")
    return device


def _command_display(argv: list[str]) -> str:
    return shlex.join(str(arg) for arg in argv)


def _build_pick_command(data: dict, task_id: str) -> list[str]:
    return [
        sys.executable,
        str(_PROJECT_ROOT / "pnsn" / "picker.py"),
        "-i", _require_text(data, "input_dir", "Input directory"),
        "-o", str(_PROJECT_ROOT / "web_app" / "outputs" / task_id),
        "-m", _optional_text(data, "model", "pnsn/pickers/pnsn.v3.jit"),
        "-d", _validated_device(data),
    ]


def _build_association_command(data: dict, task_id: str) -> list[str]:
    method_scripts = {
        "fastlink": "pnsn/fastlinker.py",
        "real": "pnsn/reallinker.py",
        "gamma": "pnsn/gammalink.py",
    }
    method = _optional_text(data, "method", "fastlink")
    script = method_scripts.get(method)
    if not script:
        raise ValueError("Invalid method; expected fastlink, real, or gamma")
    argv = [
        sys.executable,
        str(_PROJECT_ROOT / script),
        "-i", _require_text(data, "input_file", "Input file"),
        "-o", str(_PROJECT_ROOT / "web_app" / "outputs" / f"{task_id}.txt"),
        "-s", _require_text(data, "station_file", "Station file"),
    ]
    if method == "fastlink":
        argv.extend(["-d", _validated_device(data)])
    return argv


def _build_polarity_command(data: dict, task_id: str) -> list[str]:
    return [
        sys.executable,
        str(_PROJECT_ROOT / "seismic_cli.py"),
        "polarity",
        "-i", _require_text(data, "input_file", "Input file"),
        "-w", _require_text(data, "waveform_dir", "Waveform directory"),
        "-o", str(_PROJECT_ROOT / "web_app" / "outputs" / f"{task_id}_polarity.txt"),
        "--model", _optional_text(data, "model", "pnsn/pickers/polar.onnx"),
        "--min-confidence", str(data.get("min_confidence", 0.5)),
        "--phase", _optional_text(data, "phase", "Pg"),
    ]


def run_task(task_id, command, task_type, cwd=None):
    """Run a seismic processing task in background"""
    try:
        tasks[task_id]['status'] = 'running'
        tasks[task_id]['start_time'] = datetime.now().isoformat()

        argv = list(command) if isinstance(command, (list, tuple)) else shlex.split(str(command))
        result = subprocess.run(
            argv,
            shell=False,
            capture_output=True,
            text=True,
            timeout=3600,  # 1 hour timeout
            cwd=cwd or os.getcwd()
        )

        tasks[task_id]['end_time'] = datetime.now().isoformat()
        tasks[task_id]['returncode'] = result.returncode
        tasks[task_id]['stdout'] = result.stdout[-5000:] if result.stdout else ""  # Last 5000 chars
        tasks[task_id]['stderr'] = result.stderr[-5000:] if result.stderr else ""

        if result.returncode == 0:
            tasks[task_id]['status'] = 'completed'
        else:
            tasks[task_id]['status'] = 'failed'

    except subprocess.TimeoutExpired:
        tasks[task_id]['status'] = 'timeout'
        tasks[task_id]['stderr'] = "Task timed out (1 hour limit)"
    except Exception as e:
        tasks[task_id]['status'] = 'error'
        tasks[task_id]['stderr'] = str(e)


@bp.route('/api/tasks', methods=['GET'])
def list_tasks():
    """List all tasks"""
    return jsonify({
        'tasks': {k: {kk: vv for kk, vv in v.items() if kk not in ['stdout', 'stderr']}
                  for k, v in tasks.items()}
    })


@bp.route('/api/task/<task_id>', methods=['GET'])
def get_task(task_id):
    """Get task status and results"""
    if task_id not in tasks:
        return jsonify({'error': 'Task not found'}), 404

    task = tasks[task_id].copy()
    task['logs'] = {
        'stdout': task.get('stdout', ''),
        'stderr': task.get('stderr', '')
    }
    task.pop('stdout', None)

    return jsonify(task)


@bp.route('/api/chat_picks/<task_id>', methods=['GET'])
def get_chat_picks(task_id):
    """Poll pick task status; returns parsed picks when done."""
    if task_id not in tasks:
        return jsonify({'error': 'Task not found'}), 404
    task = tasks[task_id]
    status = task.get('status', 'running')
    if status == 'running':
        return jsonify({'status': 'running'})
    if status in ('error', 'failed', 'timeout'):
        return jsonify({'status': status, 'error': task.get('stderr', '')})
    # Parse picks file
    picks_file = task.get('picks_file', '')
    picks = []
    if os.path.exists(picks_file):
        with open(picks_file, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith('#'):
                    continue
                parts = line.split(',')
                if len(parts) < 4:
                    continue
                try:
                    picks.append({
                        'phase': parts[0],
                        'time_s': float(parts[1]),
                        'confidence': float(parts[2]),
                        'abs_time': parts[3],
                        'snr': float(parts[4]) if len(parts) > 4 else 0.0,
                        'station': parts[6] if len(parts) > 6 else '',
                        'polarity': parts[7] if len(parts) > 7 else 'N',
                    })
                except (ValueError, IndexError):
                    continue
    return jsonify({'status': 'completed', 'picks': picks, 'n_picks': len(picks)})


@bp.route('/api/pick', methods=['POST'])
def submit_picking():
    """Submit phase picking job"""
    data = request.json or {}

    task_id = f"pick_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{os.getpid()}"

    try:
        cmd = _build_pick_command(data, task_id)
    except ValueError as exc:
        return jsonify({'error': str(exc)}), 400

    # Initialize task
    tasks[task_id] = {
        'id': task_id,
        'type': 'phase_picking',
        'status': 'queued',
        'command': _command_display(cmd),
        'parameters': data,
        'created_at': datetime.now().isoformat()
    }

    # Start task in background
    thread = threading.Thread(target=run_task, args=(task_id, cmd, 'picking'))
    thread.daemon = True
    thread.start()

    return jsonify({'task_id': task_id, 'message': 'Task submitted'})


@bp.route('/api/associate', methods=['POST'])
def submit_association():
    """Submit phase association job"""
    data = request.json or {}

    task_id = f"assoc_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{os.getpid()}"

    try:
        cmd = _build_association_command(data, task_id)
    except ValueError as exc:
        return jsonify({'error': str(exc)}), 400

    # Initialize task
    tasks[task_id] = {
        'id': task_id,
        'type': 'phase_association',
        'status': 'queued',
        'command': _command_display(cmd),
        'parameters': data,
        'created_at': datetime.now().isoformat()
    }

    # Start task in background
    thread = threading.Thread(target=run_task, args=(task_id, cmd, 'association'))
    thread.daemon = True
    thread.start()

    return jsonify({'task_id': task_id, 'message': 'Task submitted'})


@bp.route('/api/polarity', methods=['POST'])
def submit_polarity():
    """Submit polarity analysis job"""
    data = request.json or {}

    task_id = f"polar_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{os.getpid()}"

    try:
        cmd = _build_polarity_command(data, task_id)
    except ValueError as exc:
        return jsonify({'error': str(exc)}), 400

    # Initialize task
    tasks[task_id] = {
        'id': task_id,
        'type': 'polarity_analysis',
        'status': 'queued',
        'command': _command_display(cmd),
        'parameters': data,
        'created_at': datetime.now().isoformat()
    }

    # Start task in background
    thread = threading.Thread(target=run_task, args=(task_id, cmd, 'polarity'))
    thread.daemon = True
    thread.start()

    return jsonify({'task_id': task_id, 'message': 'Task submitted'})


@bp.route('/api/output/<filename>', methods=['GET'])
def download_output(filename):
    """Download output file"""
    filepath = os.path.join('web_app/outputs', filename)
    if os.path.exists(filepath):
        return send_file(filepath, as_attachment=True)
    return jsonify({'error': 'File not found'}), 404


# ── Literature Loop Agent ───────────────────────────────────────────────────

def _lit_gc():
    cutoff = _time.time() - 1800  # 30-min TTL (reports are larger than code results)
    for k in [k for k, v in _lit_jobs.items() if v.get("ts", 0) < cutoff]:
        _lit_jobs.pop(k, None)


@bp.route('/api/literature_loop', methods=['POST'])
def literature_loop():
    """Start an async literature-loop interpretation job."""
    data          = request.json or {}
    question      = (data.get("question") or "").strip()
    study_area    = (data.get("study_area") or "").strip()
    max_iters     = int(data.get("max_iterations", 3))
    top_k         = int(data.get("rag_top_k", 8))

    if not question:
        return jsonify({"ok": False, "error": "question is required"}), 400

    _lit_gc()
    job_id = "lit_" + _uuid.uuid4().hex[:10]
    _lit_jobs[job_id] = {
        "status":   "running",
        "progress": [],
        "result":   None,
        "error":    None,
        "ts":       _time.time(),
    }

    def _run():
        try:
            import sys as _sys
            import os as _os
            _root = _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__)))
            if _root not in _sys.path:
                _sys.path.insert(0, _root)
            from sage_agents import LiteratureLoopAgent

            def _prog(d):
                phase = d.get("phase", "")
                msg   = d.get("message") or d.get("msg", "")
                _lit_jobs[job_id]["progress"].append(
                    {"phase": phase, "message": msg, "ts": _time.time()}
                )

            agent  = LiteratureLoopAgent(llm_cfg=get_llm_config(), top_k=top_k)
            result = agent.run(question, study_area, max_iterations=max_iters,
                               on_progress=_prog)
            _lit_jobs[job_id]["status"] = "done"
            _lit_jobs[job_id]["result"] = agent.result_to_dict(result)
        except Exception as exc:
            _lit_jobs[job_id]["status"] = "error"
            _lit_jobs[job_id]["error"]  = str(exc)

    threading.Thread(target=_run, daemon=True).start()
    return jsonify({"ok": True, "job_id": job_id})


@bp.route('/api/literature_loop/poll/<job_id>', methods=['GET'])
def literature_loop_poll(job_id):
    """Poll for literature-loop job status."""
    job = _lit_jobs.get(job_id)
    if not job:
        return jsonify({"ok": False, "error": "Job not found"}), 404
    return jsonify({
        "ok":       True,
        "status":   job["status"],
        "progress": job["progress"],
        "result":   job["result"],
        "error":    job["error"],
    })


# ── Scientific Analysis Agent alpha ─────────────────────────────────────────

def _science_agent_gc():
    """Discard scientific analysis jobs older than 2 hours."""
    cutoff = _time.time() - 7200
    for k in [k for k, v in _science_agent_jobs.items() if v.get("ts", 0) < cutoff]:
        _science_agent_jobs.pop(k, None)


@bp.route('/api/science_analysis_agent', methods=['POST'])
def science_analysis_agent():
    """Start an async autonomous scientific analysis job."""
    data = request.json or {}
    question = (data.get("question") or "").strip()
    data_description = (data.get("data_description") or "").strip()
    if not question and not data_description:
        return jsonify({"ok": False, "error": "question or data_description is required"}), 400

    _science_agent_gc()
    job_id = "sci_" + _uuid.uuid4().hex[:10]
    session_id = _clean_conversation_id(data.get("session_id") or "default_science")
    workspace_root = _science_resolve_workspace_root(data.get("workspace_root"), session_id)
    output_cfg = Path(data.get("output_dir") or "outputs/science_analysis_agent").expanduser()
    if output_cfg.is_absolute():
        try:
            output_cfg.resolve().relative_to(workspace_root.resolve())
            effective_output_dir = output_cfg / session_id / job_id
        except Exception:
            effective_output_dir = workspace_root / "outputs" / "science_analysis_agent" / session_id / job_id
    else:
        effective_output_dir = workspace_root / output_cfg / session_id / job_id
    _science_agent_jobs[job_id] = {
        "status": "running",
        "progress": [],
        "guidance": [],
        "result": None,
        "error": None,
        "ts": _time.time(),
        "session_id": session_id,
        "workspace_root": str(workspace_root),
        "output_dir": str(effective_output_dir),
    }

    def _run():
        try:
            import sys as _sys
            import os as _os
            _root = _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__)))
            if _root not in _sys.path:
                _sys.path.insert(0, _root)
            _agent_backend = "seismo_agent"

            ws_cfg = get_workspace_config()
            authorized_roots = []
            if ws_cfg.get("enabled"):
                authorized_roots.extend(ws_cfg.get("paths") or [])
            authorized_roots.extend(data.get("authorized_roots") or [])
            authorized_roots = [str(p).strip() for p in authorized_roots if str(p).strip()]

            workspace_root_str = str(workspace_root)
            literature_root = data.get("literature_root") or ""

            def _prog(d):
                if isinstance(d, str):
                    phase, msg = "agent", d
                else:
                    phase = d.get("phase", "")
                    msg = d.get("message") or d.get("msg", "")
                _science_agent_jobs[job_id]["progress"].append(
                    {"phase": phase, "message": msg, "ts": _time.time()}
                )

            def _runtime_guidance() -> str:
                items = _science_agent_jobs.get(job_id, {}).get("guidance") or []
                return "\n".join(f"- {g.get('message', '')}" for g in items if g.get("message"))

            file_profiles = []
            file_role_summary = ""
            try:
                root_path = Path(workspace_root_str).expanduser()
                _prog({"phase": "file_profile", "message": f"Scanning project files in {root_path}"})
                file_profiles = _science_build_file_profiles(root_path)
                _prog({"phase": "file_profile", "message": f"Profiled {len(file_profiles)} project files"})
                if file_profiles:
                    file_role_summary = _science_llm_role_summary(file_profiles, data_description)
                    _prog({"phase": "file_profile", "message": "Prepared LLM-assisted file-role assessment"})
            except Exception as exc:
                file_profiles = []
                file_role_summary = f"[file profiling failed: {exc}]"
            literature_sources = []
            seen_lit_keys = set()
            for p in file_profiles:
                if not (p.get("suffix") == ".pdf" or p.get("role_hint") == "reference_paper_or_report"):
                    continue
                abs_path = str(p.get("abs_path") or "")
                if not abs_path:
                    continue
                rel_path = str(p.get("path") or "").replace("\\", "/")
                # article/ contains templates and compiled manuscript drafts, not
                # source literature. Loading those as papers pollutes the evidence
                # context and duplicates the agent's own prior outputs.
                if rel_path.startswith("article/"):
                    continue
                key = (Path(abs_path).name.lower(), int(p.get("size") or 0))
                if key in seen_lit_keys:
                    continue
                seen_lit_keys.add(key)
                literature_sources.append(abs_path)
            if literature_sources:
                _prog({
                    "phase": "literature",
                    "message": f"Detected {len(literature_sources)} local PDF literature files",
                })

            profile = get_user_profile_context(max_chars=2500)
            project_context = (data.get("project_context") or "").strip()
            prompt_parts = [
                "你是 Scientific Analysis Agent alpha，不是参数优化专用助手。",
                "目标是把用户提供的数据说明、工作目录数据、本地/在线文献和知识库证据，转化为可复现的科学分析、报告和论文草稿。",
                f"项目根目录：{workspace_root_str}",
                "编程执行时当前工作目录就是项目根目录；请使用项目相对路径（如 data/xxx、literature/xxx、docs/xxx）访问文件。",
                f"所有输出必须写入：{effective_output_dir}",
                "",
                "===== User research request =====",
                question or "请根据数据说明和工作目录自主开展科学分析。",
            ]
            if data_description:
                prompt_parts += ["", "===== Data description supplied by user =====", data_description]
            if file_profiles or file_role_summary:
                prompt_parts += [
                    "",
                    "===== Initial project file profiles =====",
                    _science_profiles_to_prompt(file_profiles, file_role_summary),
                ]
            prompt_parts += [
                "",
                "===== Required autonomous workflow =====",
                "1. 数据研究：根据文件画像、数据说明、文献和必要的 web search，先判断这些数据可以支持哪些研究方向；用户指定方向时优先服从用户方向。",
                "2. 科学问题规划：基于数据可用性和文献背景，提出可验证的科学问题、假设、反证路径和缺失信息。",
                "3. 多技能/RAG 编排：可同时调用或交叉调用多个 SKILL（如 deep-research、academic-paper、academic-paper-reviewer、领域绘图/数据处理技能）和知识库 RAG；先选择技能组合，再开展分析。",
                "4. 多模态证据：如果启用了多模态且当前模型支持图像/表格解析，应分析上传图片、论文图件和表格并提取定量证据；如果模型不支持，必须明确警示并退回文本/数值证据。",
                "5. 图件与统计规划：让 LLM 先规划论文需要哪些图件、统计量、表格和中间产物，再进入编程；主文图件默认不超过 3 张、主表默认不超过 2 张，必须服务于机制假设检验。",
                "6. Coding Agent 执行：自己编程解析数据格式、整理字段、做 mini test、统计和绘图；出错后不要停，必须让 LLM 根据错误自动诊断、改方案、重试。",
                "7. 证据综合：写论文前必须把数据统计、图件、表格、本地论文/RAG 和 web search 整合成 claim-evidence-warrant 矩阵；每个科学结论都要列出支持证据、反证路径和缺失信息。",
                "8. 结论驱动图表复审：形成初步结论后，必须反向判断哪些图表应保留为主文、哪些降为补充/QC、哪些缺口需要补图补表，并据此迭代一次论证。",
                "9. 论文撰写：根据生成图件、统计结果、给定论文和 web search 证据撰写 Markdown 论文草稿，并用 Markdown 图片语法嵌入图件；Results 是文章核心，必须围绕新的机制性科学结论，而不是围绕数据质量或图件说明。",
                "10. 三审稿人循环：模拟 3 个严格审稿人分别审查机制创新、数据/统计可重复性、文献证据与写作；根据意见修订，直到三位均为小修/接收或达到轮次上限。",
                "11. 交互式迭代：如果用户后续提出修改、补充或新假设，基于上一轮结果继续研究，不要从零开始。",
                "12. 所有读写、脚本、图表、报告、LaTeX 和临时文件都必须限制在用户指定的项目工作目录内；不要写到项目目录之外。",
                "13. 必须使用上面的文件画像判断文件作用：numeric_text_data/structured_data/waveform_data 都应视为候选数据；reference_paper_or_report 视为参考文献；data_description_or_project_notes 视为数据说明。不要因为文件扩展名是 .txt/.doc/.pdf 就忽略它。",
                "14. 参数优化复用：如果项目中存在 `science_analysis_inputs/parameter_optimization/`、`parameter_optimization_run.md/json`、`best_parameters.json`、`optimization_history.csv` 或 `optimization_report.md`，必须把它们视为实验/方法证据；可以据此撰写参数优化方法、结果、消融/敏感性分析和论文补充材料，但必须标注 dry-run/失败/缺失信息。",
                "15. 所有结论必须有来源依据：数据文件、统计输出、图件、参数优化运行记录、RAG chunk、web 文献或工具输出。证据不足就明确说明。",
                "16. 不要把中间猜测写成事实；报告中保留“已验证/待验证/缺失信息”状态。",
                "17. 这不是数据质量分析页面：不要把质量分级、字段清单、参数直方图、基础计数作为主线；这些只能进入补充 QC 文件。主线必须围绕科学问题，例如断层几何、弱层/低速体、应力转移、分段破裂、流体或应变释放机制。",
            ]
            if project_context:
                prompt_parts += ["", "===== Project shared context =====", project_context[:16000]]
            if profile:
                prompt_parts += ["", "===== Long-term user profile (soft context; do not mention unless useful) =====", profile]

            study_area = (data.get("study_area") or "scientific analysis").strip()
            _prog({"phase": "start", "message": f"Scientific analysis workspace: {workspace_root_str}"})
            _prog({"phase": "backend", "message": f"Using {_agent_backend}"})
            from seismo_agent import SeismoAgent
            agent = SeismoAgent(llm_config=get_llm_config(), project_root=workspace_root_str, mode="autonomous")
            fallback = agent.run(
                "\n".join(prompt_parts),
                paper_source=literature_sources,
                output_dir=str(effective_output_dir),
                progress_cb=_prog,
                guidance_provider=_runtime_guidance,
                max_steps=int(data.get("max_iterations", 4)),
                max_followup_rounds=int(data.get("max_followup_rounds", 3)),
                max_review_rounds=int(data.get("max_review_rounds", 3)),
                produce_latex=bool(data.get("produce_latex", True)),
                use_web_search=bool(data.get("allow_web_search", True)),
            )
            result = {
                "final_report": fallback.get("summary", ""),
                "generated_figures": fallback.get("figures", []),
                "markdown_paper": fallback.get("markdown_paper", ""),
                "markdown_paper_path": fallback.get("markdown_paper_path", ""),
                "latex_path": fallback.get("latex_path", ""),
                "latex_bib_path": fallback.get("latex_bib_path", ""),
                "latex_pdf": fallback.get("latex_pdf", ""),
                "latex_paper": fallback.get("latex_paper", ""),
                "tool_log": [],
                "scientific_questions": fallback.get("scientific_questions", []),
                "statistical_results": fallback.get("statistical_results") or [
                    {"title": "Agent summary", "content": fallback.get("summary", "")}
                ],
                "table_artifacts": fallback.get("table_artifacts", []),
                "paper_artifact_plan": fallback.get("paper_artifact_plan", ""),
                "scientific_evidence_synthesis": fallback.get("scientific_evidence_synthesis", ""),
                "artifact_refinement_plan": fallback.get("artifact_refinement_plan", ""),
                "peer_review_reports": fallback.get("peer_review_reports", []),
                "missing_information": [] if fallback.get("success") else ["部分步骤失败，请查看运行日志。"],
                "_fallback_backend": "seismo_agent",
                "_raw_result": fallback,
            }
            if isinstance(result, dict):
                result.setdefault("_run_output_dir", str(effective_output_dir))
                result.setdefault("_session_id", session_id)
                result.setdefault("_agent_kind", "science_analysis")
                result.setdefault("_workspace_root", workspace_root_str)
                _science_finalize_result_artifacts(result, effective_output_dir, _prog)
            _science_agent_jobs[job_id]["status"] = "done"
            _science_agent_jobs[job_id]["result"] = result
        except Exception as exc:
            _science_agent_jobs[job_id]["status"] = "error"
            _science_agent_jobs[job_id]["error"] = str(exc)

    threading.Thread(target=_run, daemon=True).start()
    return jsonify({"ok": True, "job_id": job_id, "output_dir": str(effective_output_dir)})


@bp.route('/api/science_analysis_agent/guidance', methods=['POST'])
def science_analysis_agent_guidance():
    data = request.json or {}
    job_id = (data.get("job_id") or "").strip()
    message = (data.get("message") or "").strip()
    if not job_id or not message:
        return jsonify({"ok": False, "error": "job_id and message are required"}), 400
    job = _science_agent_jobs.get(job_id)
    if not job:
        return jsonify({"ok": False, "error": "Job not found"}), 404
    item = {"message": message, "ts": _time.time()}
    job.setdefault("guidance", []).append(item)
    job.setdefault("progress", []).append({"phase": "guidance", "message": message, "ts": _time.time()})
    return jsonify({"ok": True, "guidance": item})


@bp.route('/api/science_analysis_agent/poll/<job_id>', methods=['GET'])
def science_analysis_agent_poll(job_id):
    job = _science_agent_jobs.get(job_id)
    if not job:
        return jsonify({"ok": False, "error": "Job not found"}), 404
    return jsonify({
        "ok": True,
        "status": job["status"],
        "progress": job["progress"],
        "guidance": job.get("guidance", []),
        "result": job["result"],
        "error": job["error"],
        "workspace_root": job.get("workspace_root", ""),
        "output_dir": job.get("output_dir", ""),
    })


_SCI_ALLOWED_EXTS = {
    ".pdf", ".png", ".jpg", ".jpeg", ".svg", ".tif", ".tiff",
    ".csv", ".txt", ".md", ".json",
    ".yaml", ".yml", ".bib", ".dat",
    ".sac", ".mseed", ".xml",
    ".xlsx", ".xls", ".tsv", ".parquet", ".h5", ".hdf5", ".nc",
    ".py", ".ipynb", ".sh", ".tex", ".doc", ".docx", ".ppt", ".pptx",
}


@bp.route('/api/science_analysis_agent/upload', methods=['POST'])
def science_analysis_agent_upload():
    """Upload a research data/literature file into the science workspace."""
    if 'file' not in request.files:
        return jsonify({"ok": False, "error": "No file provided"}), 400
    f = request.files['file']
    session_id = _clean_conversation_id(request.form.get('session_id') or 'default_science')
    orig_name = Path(f.filename).name if f.filename else 'upload'
    ext = Path(orig_name).suffix.lower()
    if ext not in _SCI_ALLOWED_EXTS:
        return jsonify({"ok": False, "error": f"File type '{ext}' not allowed"}), 400

    workspace_root = request.form.get('workspace_root') or ""
    ws_dir = _science_resolve_workspace_root(workspace_root, session_id)
    if ext == ".pdf":
        sub = "literature"
    elif ext in {".png", ".jpg", ".jpeg", ".svg", ".tif", ".tiff"}:
        sub = "figures"
    elif ext in {".csv", ".tsv", ".xlsx", ".xls", ".parquet", ".h5", ".hdf5", ".nc", ".dat", ".json"}:
        sub = "data"
    elif ext in {".py", ".ipynb", ".sh"}:
        sub = "code"
    else:
        sub = "docs"
    dest_dir = ws_dir / sub
    dest_dir.mkdir(parents=True, exist_ok=True)
    dest = dest_dir / orig_name
    f.save(str(dest))
    preview = _science_read_file_preview(dest, max_chars=1600)
    role_hint = _science_guess_file_role(dest, preview)
    original_sub = sub
    if role_hint in {"numeric_text_data", "structured_data", "waveform_data"}:
        sub = "data"
    elif role_hint == "data_description_or_project_notes":
        sub = "docs"
    if sub != original_sub:
        new_dir = ws_dir / sub
        new_dir.mkdir(parents=True, exist_ok=True)
        new_dest = new_dir / orig_name
        try:
            dest.replace(new_dest)
            dest = new_dest
        except Exception:
            pass
    return jsonify({
        "ok": True,
        "path": str(dest),
        "session_workspace": str(ws_dir),
        "file_type": sub,
        "role_hint": role_hint,
        "preview": preview[:500],
    })


@bp.route('/api/science_analysis_agent/artifact', methods=['GET'])
def science_analysis_agent_artifact():
    """Serve a generated or uploaded science-analysis artifact."""
    artifact_path = request.args.get('path', '').strip()
    if not artifact_path:
        return jsonify({"error": "path required"}), 400
    p = Path(artifact_path).expanduser()
    if not p.exists() or not p.is_file():
        return jsonify({"error": "File not found"}), 404
    allowed = {'.png', '.jpg', '.jpeg', '.svg', '.html', '.pdf', '.txt', '.md', '.csv', '.json'}
    if p.suffix.lower() not in allowed:
        return jsonify({"error": "Unsupported file type"}), 400
    resolved = p.resolve()
    allowed_roots = [Path(__file__).parent.parent.resolve(), SCIENCE_WORKSPACE_ROOT.resolve()]
    for job in _science_agent_jobs.values():
        for key in ("workspace_root", "output_dir"):
            val = job.get(key)
            if val:
                try:
                    allowed_roots.append(Path(val).expanduser().resolve())
                except Exception:
                    pass
    allowed = False
    for root in allowed_roots:
        try:
            resolved.relative_to(root)
            allowed = True
            break
        except ValueError:
            continue
    if not allowed:
        return jsonify({"error": "Access denied"}), 403
    if p.suffix.lower() == '.svg':
        return send_file(str(resolved), mimetype='image/svg+xml')
    return send_file(str(resolved))


@bp.route('/api/science_analysis_agent/export_pdf', methods=['POST'])
def science_analysis_agent_export_pdf():
    """Export an already-generated Markdown science paper to HTML/PDF."""
    data = request.json or {}
    session_id = _clean_conversation_id(data.get("session_id") or "default_science")
    workspace_root = _science_resolve_workspace_root(data.get("workspace_root"), session_id)
    output_raw = str(data.get("output_dir") or data.get("_run_output_dir") or "").strip()
    if output_raw:
        out = Path(output_raw).expanduser()
        if not out.is_absolute():
            out = workspace_root / out
    else:
        out = workspace_root / "outputs" / "science_analysis_agent" / session_id / "manual_export"
    try:
        out_resolved = out.resolve()
        ws_resolved = workspace_root.resolve()
        if ws_resolved not in [out_resolved, *out_resolved.parents]:
            out = ws_resolved / "outputs" / "science_analysis_agent" / session_id / "manual_export"
    except Exception:
        out = workspace_root / "outputs" / "science_analysis_agent" / session_id / "manual_export"
    result = {
        "markdown_paper": data.get("markdown_paper") or "",
        "markdown_paper_path": data.get("markdown_paper_path") or "",
        "generated_figures": data.get("generated_figures") or [],
    }
    logs = []
    _science_finalize_result_artifacts(result, out, lambda d: logs.append(d))
    if not (result.get("paper_pdf") or result.get("paper_html")):
        return jsonify({"ok": False, "error": result.get("paper_pdf_error") or result.get("paper_html_error") or "No paper content to export"}), 400
    _science_agent_jobs[f"export_{session_id}"] = {
        "status": "done",
        "progress": logs,
        "guidance": [],
        "result": result,
        "error": None,
        "ts": _time.time(),
        "session_id": session_id,
        "workspace_root": str(workspace_root),
        "output_dir": str(out),
    }
    return jsonify({"ok": True, "result": result, "logs": logs, "output_dir": str(out)})


# ── 聊天界面临时文档上传 ────────────────────────────────────────────────────

@bp.route('/api/chat/upload', methods=['POST'])
def chat_upload_pdf():
    """Upload a PDF for the current chat session (temporary RAG, not persisted)."""
    if 'file' not in request.files:
        return jsonify({"ok": False, "error": "No file"}), 400

    f = request.files['file']
    session_id = _clean_conversation_id(request.form.get('session_id', 'default'))
    project_id = _clean_conversation_id(request.form.get('project_id', ''))

    upload_id = _uuid.uuid4().hex[:10]
    try:
        tmp_path, _safe_name = _safe_chat_pdf_upload_path(f.filename, session_id, upload_id)
    except ValueError as exc:
        return jsonify({"ok": False, "error": str(exc)}), 400
    f.save(str(tmp_path))

    try:
        # Extract text (no BGE-M3, just raw text for session context)
        pages, chunks = _extract_session_pdf_chunks(tmp_path, upload_id, f.filename)
        for chunk in chunks:
            chunk.setdefault("doc_name", f.filename)
            chunk.setdefault("upload_id", upload_id)

        if session_id not in _session_docs:
            _session_docs[session_id] = {"chunks": [], "doc_names": [], "files": {}}

        _session_docs[session_id]["chunks"].extend(chunks)
        _session_docs[session_id]["doc_names"].append(f.filename)
        _session_docs[session_id].setdefault("files", {})[upload_id] = {
            "name": f.filename,
            "path": str(tmp_path),
            "n_pages": len(pages),
            "n_chunks": len(chunks),
            "permanent": False,
        }

        if project_id:
            project_key = f"project_{project_id}"
            if project_key not in _session_docs:
                _session_docs[project_key] = {"chunks": [], "doc_names": [], "files": {}}
            _session_docs[project_key]["chunks"].extend(chunks)
            if f.filename not in _session_docs[project_key]["doc_names"]:
                _session_docs[project_key]["doc_names"].append(f.filename)
            _session_docs[project_key].setdefault("files", {})[upload_id] = {
                "name": f.filename,
                "path": str(tmp_path),
                "n_pages": len(pages),
                "n_chunks": len(chunks),
                "permanent": False,
            }

        return jsonify({
            "ok": True,
            "doc_name": f.filename,
            "upload_id": upload_id,
            "n_pages":  len(pages),
            "n_chunks": len(chunks),
            "session_id": session_id,
            "project_id": project_id,
        })
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@bp.route('/api/chat/promote_doc', methods=['POST'])
def chat_promote_doc():
    """Persist a temporary chat PDF into the permanent knowledge base."""
    data = request.get_json(silent=True) or {}
    session_id = _clean_conversation_id(data.get("session_id", "default"))
    upload_id = data.get("upload_id", "")
    info = _session_docs.get(session_id, {}).get("files", {}).get(upload_id)
    if not info:
        return jsonify({"ok": False, "error": "Temporary document not found"}), 404

    path = Path(info.get("path", ""))
    if not path.exists():
        return jsonify({"ok": False, "error": "Temporary file has expired"}), 404

    try:
        kb = get_kb_instance()
        if not kb:
            return jsonify({"ok": False, "error": "Knowledge base unavailable"}), 500
        logs = []
        meta = kb.add_pdf(str(path), progress_cb=lambda m: logs.append(m), source_type="upload")
        info["permanent"] = True
        return jsonify({
            "ok": True,
            "doc_id": meta.doc_id,
            "doc_name": meta.doc_name,
            "n_chunks": meta.n_chunks,
            "logs": logs[-20:],
        })
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 500


def _remove_doc_from_session(session_key: str, upload_id: str = "", doc_name: str = "") -> bool:
    session = _session_docs.get(session_key)
    if not session:
        return False

    files = session.setdefault("files", {})
    removed_infos = []
    if upload_id and upload_id in files:
        removed_infos.append(files.pop(upload_id))
    elif doc_name:
        for fid, info in list(files.items()):
            if info.get("name") == doc_name:
                removed_infos.append(files.pop(fid))

    removed_upload_ids = {upload_id} if upload_id else set()
    removed_names = {doc_name} if doc_name else set()
    for info in removed_infos:
        if info.get("name"):
            removed_names.add(info.get("name"))
        try:
            if not info.get("permanent"):
                Path(info.get("path", "")).unlink(missing_ok=True)
        except Exception:
            pass

    def _should_remove_chunk(chunk: dict) -> bool:
        return (
            (upload_id and chunk.get("upload_id") in removed_upload_ids)
            or (doc_name and chunk.get("doc_name") in removed_names)
        )

    before = len(session.get("chunks") or [])
    session["chunks"] = [
        c for c in (session.get("chunks") or [])
        if not _should_remove_chunk(c)
    ]

    remaining_names = {
        info.get("name")
        for info in files.values()
        if info.get("name")
    } | {
        c.get("doc_name")
        for c in session.get("chunks", [])
        if c.get("doc_name")
    }
    if remaining_names:
        session["doc_names"] = [
            n for n in (session.get("doc_names") or [])
            if n in remaining_names
        ]
    else:
        session["doc_names"] = []

    if not session.get("chunks") and not session.get("files") and not session.get("doc_names"):
        _session_docs.pop(session_key, None)

    return bool(removed_infos) or len(session.get("chunks") or []) != before


@bp.route('/api/chat/remove_session_doc', methods=['POST'])
def chat_remove_session_doc():
    data = request.get_json(silent=True) or {}
    sid = _clean_conversation_id(data.get('session_id', 'default'))
    project_id = _clean_conversation_id(data.get('project_id', ''))
    upload_id = str(data.get("upload_id") or "").strip()
    doc_name = str(data.get("doc_name") or "").strip()
    if not upload_id and not doc_name:
        return jsonify({"ok": False, "error": "Missing document identifier"}), 400

    removed = _remove_doc_from_session(sid, upload_id=upload_id, doc_name=doc_name)
    if project_id:
        removed = _remove_doc_from_session(f"project_{project_id}", upload_id=upload_id, doc_name=doc_name) or removed
    session = _session_docs.get(sid, {})
    return jsonify({
        "ok": True,
        "removed": removed,
        "doc_names": session.get("doc_names", []),
    })


@bp.route('/api/chat/clear_session', methods=['POST'])
def chat_clear_session():
    data = request.get_json(silent=True) or {}
    sid = _clean_conversation_id(data.get('session_id', 'default'))
    for info in _session_docs.get(sid, {}).get("files", {}).values():
        try:
            Path(info.get("path", "")).unlink(missing_ok=True)
        except Exception:
            pass
    _session_docs.pop(sid, None)
    return jsonify({"ok": True})


# ── RAG 增强对话 ──────────────────────────────────────────────────────────────

@bp.route('/api/chat/rag', methods=['POST'])
def chat_rag():
    """RAG-aware chat endpoint."""
    data       = request.json or {}
    user_msg   = data.get("message", "").strip()
    session_id = _clean_conversation_id(data.get("session_id", "default"))
    mode       = data.get("mode", "rag")   # "rag" | "paper_read"

    if not user_msg:
        return jsonify({"ok": False, "error": "Empty message"}), 400

    llm_cfg = get_llm_config()
    if not llm_cfg.get("api_base"):
        return jsonify({
            "ok": True,
            "response": (
                "当前没有可用的 LLM 后端。\n"
                "请在 **LLM 设置** 页面配置后端（Ollama / 在线 API）。"
            ),
            "sources": [],
        })

    # ── 检索上下文 ──────────────────────────────────────────────────────────
    context_parts = []
    sources = []

    # 0. 工作目录文件系统上下文（用户授权后注入）
    workspace_path = data.get("workspace", "")
    if workspace_path:
        ws_ctx = inject_workspace_context(user_msg, workspace_path)
        if ws_ctx:
            context_parts.append("===== 本地文件系统 =====\n" + ws_ctx)

    # 1. 会话文档（临时上传）
    session = _session_docs.get(session_id, {})
    if session.get("chunks"):
        context_parts.append(_RAG_COMPLETE_ONLY_RULE)
        # 简单 TF-IDF 式关键词匹配（无需 GPU）
        query_words = set(user_msg.lower().split())
        scored = []
        for c in session["chunks"]:
            words = set(c["text"].lower().split())
            score = len(query_words & words) / (len(query_words) + 1)
            scored.append((score, c))
        scored.sort(key=lambda x: x[0], reverse=True)
        top = scored[:4]
        for score, c in top:
            if score > 0 or mode == "paper_read":
                chunk_text = _drop_incomplete_fenced_tail(c["text"])
                if not chunk_text.strip():
                    continue
                context_parts.append(
                    f"[上传文档 第{c['page']}页]\n{chunk_text}"
                )
        if session.get("doc_names"):
            sources.extend(session["doc_names"])

    # 2. 持久知识库（BGE-M3 向量检索 / TF-IDF 回退）
    try:
        kb = get_kb_instance()
        if kb and not kb.is_empty:
            kb_hits = kb.retrieve(user_msg, top_k=5, score_threshold=0.45)
            if kb_hits:
                lines = ["The following passages were retrieved from the knowledge base. "
                         "Use them only if they are directly relevant to the question:\n",
                         _RAG_COMPLETE_ONLY_RULE]
                total = 0
                for chunk, score in kb_hits:
                    chunk_text = _drop_incomplete_fenced_tail(chunk.text)
                    if not chunk_text.strip():
                        continue
                    entry = (
                        f"[Source: {chunk.doc_name}, page {chunk.page + 1}, "
                        f"relevance {score:.2f}]\n{chunk_text}\n"
                    )
                    if total + len(entry) > 2500:
                        break
                    lines.append(entry)
                    total += len(entry)
                    if chunk.doc_name not in sources:
                        sources.append(chunk.doc_name)
                context_parts.append("\n".join(lines))
    except Exception:
        pass

    # 3. seismo_skill 技能文档（按用户消息检索最相关技能，注入代码示例）
    try:
        from helpers import get_skill_loader
        sl = get_skill_loader()
        if sl is not None:
            skill_ctx, skill_rag_ctx = sl.build_skill_context_with_rag(
                user_msg, max_skill_chars=5000, max_rag_chars=3000, top_k=4
            )
            if skill_ctx:
                context_parts.append("===== 可用技能与函数示例 =====\n" + skill_ctx)
            if skill_rag_ctx:
                context_parts.append("===== 技能绑定知识库 =====\n" + skill_rag_ctx)
    except Exception:
        pass

    # ── 构建提示 ─────────────────────────────────────────────────────────────
    if mode == "paper_read":
        system = (
            "你是一位专业的地震学文献解读专家。\n"
            "请基于以下论文内容，用清晰的中文解读、总结或回答用户的问题。\n"
            "回答时请：\n"
            "1. 点明核心方法/创新点\n"
            "2. 解释关键公式或算法（必要时给出 Python 代码示例）\n"
            "3. 说明实验结果与结论\n"
            "4. 指出局限性或未来工作（如有）\n"
        )
    else:
        if context_parts:
            system = (
                "You are SAGE, an expert seismology assistant with deep knowledge of "
                "seismology and data processing.\n"
                "Relevant passages from the knowledge base are provided below. "
                "Use them to answer the question. "
                "If a passage is not directly relevant, rely on your own knowledge instead — "
                "do NOT cite or mention passages that are unrelated to the question.\n"
            )
        else:
            system = (
                "You are SAGE, an expert seismology assistant with deep knowledge of "
                "seismology and data processing.\n"
                "Answer the user's question using your own knowledge. "
                "Be concise and accurate.\n"
            )

    system += _scientific_grounding_policy(bool(data.get("enable_web_search")))
    if bool(data.get("enable_think", False)):
        system += _think_summary_policy()
    if context_parts:
        system += "\n\n===== Reference passages =====\n" + "\n\n".join(context_parts)

    messages = [
        {"role": "system", "content": system},
        {"role": "user",   "content": user_msg},
    ]

    # 加入历史（前端传入）
    history = data.get("history", [])
    if history:
        # 在 system 和最后 user 消息之间插入历史
        messages = [messages[0]] + history[-6:] + [messages[-1]]

    try:
        answer = llm_call(messages, llm_cfg, max_tokens=2000)
        return jsonify({
            "ok": True,
            "response": answer,
            "sources": list(set(sources)),
        })
    except Exception as e:
        return jsonify({
            "ok": True,
            "response": (
                f"LLM 调用失败：{e}\n\n"
                "请检查 LLM 设置页面中的后端配置是否正确。"
            ),
            "sources": [],
        })


# ── 流式 RAG 对话 ─────────────────────────────────────────────────────────────

def _openalex_abstract(item: dict) -> str:
    inv = item.get("abstract_inverted_index") or {}
    if not isinstance(inv, dict):
        return ""
    pairs = []
    for word, positions in inv.items():
        if isinstance(positions, list):
            for pos in positions:
                if isinstance(pos, int):
                    pairs.append((pos, word))
    return " ".join(w for _, w in sorted(pairs))


def _literature_web_search(query: str, *, sources=None, max_results: int = 6) -> list:
    """Small project-local scholarly search helper independent of legacy geo agent."""
    import urllib.parse
    import urllib.request
    import xml.etree.ElementTree as ET

    q = " ".join(str(query or "").split())[:500]
    if not q:
        return []
    sources = sources or _default_search_sources()
    max_results = max(1, min(int(max_results or 6), 12))
    items = []

    try:
        from config_manager import LLMConfigManager
        search_cfg = LLMConfigManager().get_search_config()
        providers = search_cfg.get("providers") or {}
    except Exception:
        providers = {}

    def _enabled(name: str) -> bool:
        meta = providers.get(name) or {}
        return bool(meta.get("enabled", True))

    def _json_get(url: str, headers=None, timeout=12):
        req = urllib.request.Request(url, headers=headers or {"User-Agent": "SeismicX/0.1"})
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            return json.loads(resp.read().decode("utf-8", errors="ignore"))

    if "openalex" in sources and _enabled("openalex"):
        try:
            url = (
                "https://api.openalex.org/works?search="
                + urllib.parse.quote(q)
                + f"&per-page={max_results}"
            )
            for item in (_json_get(url).get("results") or [])[:max_results]:
                authors = [
                    ((a.get("author") or {}).get("display_name") or "").strip()
                    for a in (item.get("authorships") or [])[:6]
                ]
                authors = [a for a in authors if a]
                loc = item.get("primary_location") or {}
                source = (loc.get("source") or {}).get("display_name") or "OpenAlex"
                doi = item.get("doi") or ""
                items.append({
                    "source": "openalex",
                    "title": item.get("title") or "Untitled",
                    "year": item.get("publication_year") or "",
                    "authors": authors,
                    "doi": doi,
                    "url": doi or item.get("id") or "",
                    "venue": source,
                    "abstract": _openalex_abstract(item),
                })
        except Exception:
            pass

    if len(items) < max_results and "semantic_scholar" in sources and _enabled("semantic_scholar"):
        try:
            api_key = (providers.get("semantic_scholar") or {}).get("api_key") or ""
            fields = "title,year,authors,abstract,url,venue,externalIds"
            url = (
                "https://api.semanticscholar.org/graph/v1/paper/search?query="
                + urllib.parse.quote(q)
                + f"&limit={max_results}&fields={urllib.parse.quote(fields)}"
            )
            headers = {"User-Agent": "SeismicX/0.1"}
            if api_key:
                headers["x-api-key"] = api_key
            for item in (_json_get(url, headers=headers).get("data") or [])[:max_results]:
                ext = item.get("externalIds") or {}
                items.append({
                    "source": "semantic_scholar",
                    "title": item.get("title") or "Untitled",
                    "year": item.get("year") or "",
                    "authors": [a.get("name", "") for a in (item.get("authors") or [])[:6] if a.get("name")],
                    "doi": ext.get("DOI") or "",
                    "url": item.get("url") or "",
                    "venue": item.get("venue") or "",
                    "abstract": item.get("abstract") or "",
                })
        except Exception:
            pass

    if len(items) < max_results and "arxiv" in sources and _enabled("arxiv"):
        try:
            url = (
                "https://export.arxiv.org/api/query?search_query=all:"
                + urllib.parse.quote(q)
                + f"&start=0&max_results={max_results}"
            )
            req = urllib.request.Request(url, headers={"User-Agent": "SeismicX/0.1"})
            with urllib.request.urlopen(req, timeout=12) as resp:
                root = ET.fromstring(resp.read())
            ns = {"a": "http://www.w3.org/2005/Atom"}
            for entry in root.findall("a:entry", ns)[:max_results]:
                title = " ".join((entry.findtext("a:title", default="", namespaces=ns) or "").split())
                summary = " ".join((entry.findtext("a:summary", default="", namespaces=ns) or "").split())
                url_text = entry.findtext("a:id", default="", namespaces=ns) or ""
                authors = [
                    (a.findtext("a:name", default="", namespaces=ns) or "").strip()
                    for a in entry.findall("a:author", ns)[:6]
                ]
                items.append({
                    "source": "arxiv",
                    "title": title or "Untitled",
                    "year": (entry.findtext("a:published", default="", namespaces=ns) or "")[:4],
                    "authors": [a for a in authors if a],
                    "doi": "",
                    "url": url_text,
                    "venue": "arXiv",
                    "abstract": summary,
                })
        except Exception:
            pass

    seen = set()
    deduped = []
    for item in items:
        key = (str(item.get("doi") or "").lower(), str(item.get("title") or "").lower())
        if key in seen:
            continue
        seen.add(key)
        deduped.append(item)
        if len(deduped) >= max_results:
            break
    return deduped


def _chat_web_search_context(data: dict, query: str):
    """Optional chat-side web/literature search context."""
    if not data.get("enable_web_search"):
        return "", []
    try:
        sources = data.get("web_search_sources") or _default_search_sources()
        papers = _literature_web_search(query, sources=sources, max_results=int(data.get("web_max_results", 6)))
        if not papers:
            return "", []
        lines = [
            "ONLINE SEARCH WAS PERFORMED. The searched records are listed below. "
            "Use them only when directly relevant. Cite records by [Web N], source, title, and URL. "
            "Do not invent claims beyond the abstracts/snippets. If these records are insufficient, say so."
        ]
        refs = []
        for i, p in enumerate(papers[:8], 1):
            title = p.get("title") or "Untitled"
            authors = ", ".join(p.get("authors", []) or [])
            year = p.get("year") or ""
            url = p.get("url") or p.get("pdf_url") or ""
            doi = p.get("doi") or ""
            abstract = (p.get("abstract") or p.get("snippet") or "")[:900]
            src = p.get("source") or "web"
            lines.append(
                f"[Web {i} | {src}] {title}\n"
                f"Authors: {authors}\nYear: {year}\nDOI: {doi}\nURL: {url}\n"
                f"Abstract/snippet: {abstract}\n"
            )
            label = f"[Web {i}] {src}: {title}"
            if year:
                label += f" ({year})"
            if url:
                label += f" — {url}"
            elif doi:
                label += f" — DOI: {doi}"
            refs.append(label)
        return "\n".join(lines), refs
    except Exception as exc:
        return f"Web search failed: {exc}", []


def _default_search_sources() -> list:
    try:
        from config_manager import LLMConfigManager
        cfg = LLMConfigManager().get_search_config()
        sources = cfg.get("default_sources") or ["openalex", "semantic_scholar"]
        providers = cfg.get("providers") or {}
        enabled = [
            s for s in sources
            if (providers.get(s, {}).get("enabled", True) or s in {"openalex", "semantic_scholar", "arxiv"})
        ]
        return enabled or ["openalex", "semantic_scholar"]
    except Exception:
        return ["openalex", "semantic_scholar"]


def _scientific_grounding_policy(web_enabled: bool = False) -> str:
    if web_enabled:
        source_rule = (
            "If online search records are present, include a short '检索来源' section that lists the Web IDs you used "
            "and cite claims with those IDs. "
        )
    else:
        source_rule = (
            "If the user asks for current literature, latest models, or literature-backed claims and no online search "
            "context is provided, explicitly say the answer may be incomplete/outdated instead of fabricating citations. "
        )
    return (
        "\n\nScientific grounding policy:\n"
        "- Do not invent papers, authors, years, URLs, DOIs, model names, benchmark numbers, parameters, or conclusions.\n"
        "- Separate source-backed facts from general background knowledge and from your own inference.\n"
        "- When evidence is missing or weak, say '目前没有足够来源支持' / 'insufficient evidence' and explain what would need to be checked.\n"
        "- Never present uncited guesses as literature facts.\n"
        f"- {source_rule}"
    )


def _think_summary_policy() -> str:
    return (
        "\n\n当启用思考模式时，请在最终回答前输出一个简短、可公开展示的推理摘要，"
        "并严格放在 <think>...</think> 标签内。摘要只写依据、检查点、计划或自检结果，"
        "不要暴露隐藏推理链。标签外只输出给用户看的最终回答。"
    )


def _build_rag_messages(data: dict):
    """
    Shared helper: build (messages, sources, llm_cfg) for both /api/chat/rag
    and its streaming twin.  Returns None for llm_cfg if backend not configured.
    """
    import json as _json
    user_msg   = data.get("message", "").strip()
    session_id = _clean_conversation_id(data.get("session_id", "default"))
    mode       = data.get("mode", "rag")

    llm_cfg = get_llm_config()

    context_parts = []
    sources = []

    workspace_path = data.get("workspace", "")
    if workspace_path:
        ws_ctx = inject_workspace_context(user_msg, workspace_path)
        if ws_ctx:
            context_parts.append("===== 本地文件系统 =====\n" + ws_ctx)

    session = _session_docs.get(session_id, {})
    project_id = _clean_conversation_id(data.get("project_id", ""))
    project_session = _session_docs.get(f"project_{project_id}", {}) if project_id else {}
    merged_chunks = list(project_session.get("chunks") or []) + list(session.get("chunks") or [])
    merged_doc_names = list(dict.fromkeys((project_session.get("doc_names") or []) + (session.get("doc_names") or [])))
    if merged_chunks:
        context_parts.append(_RAG_COMPLETE_ONLY_RULE)
        query_words = set(user_msg.lower().split())
        scored = []
        for c in merged_chunks:
            words = set(c["text"].lower().split())
            score = len(query_words & words) / (len(query_words) + 1)
            scored.append((score, c))
        scored.sort(key=lambda x: x[0], reverse=True)
        for score, c in scored[:4]:
            if score > 0 or mode == "paper_read":
                chunk_text = _drop_incomplete_fenced_tail(c["text"])
                if chunk_text.strip():
                    context_parts.append(f"[上传文档 第{c['page']}页]\n{chunk_text}")
        if merged_doc_names:
            sources.extend(merged_doc_names)

    project_context = (data.get("project_context") or "").strip()
    if project_context:
        context_parts.append("===== 项目共享上下文 =====\n" + project_context[:4000])

    web_ctx, web_sources = _chat_web_search_context(data, user_msg)
    if web_ctx:
        context_parts.append("===== Web literature/search context =====\n" + web_ctx)
    sources.extend(web_sources)

    try:
        kb = get_kb_instance()
        if kb and not kb.is_empty:
            kb_hits = kb.retrieve(user_msg, top_k=5, score_threshold=0.45)
            if kb_hits:
                lines = ["The following passages were retrieved from the knowledge base. "
                         "Use them only if they are directly relevant to the question:\n",
                         _RAG_COMPLETE_ONLY_RULE]
                total = 0
                for chunk, score in kb_hits:
                    chunk_text = _drop_incomplete_fenced_tail(chunk.text)
                    if not chunk_text.strip():
                        continue
                    entry = (f"[Source: {chunk.doc_name}, page {chunk.page + 1}, "
                             f"relevance {score:.2f}]\n{chunk_text}\n")
                    if total + len(entry) > 2500:
                        break
                    lines.append(entry)
                    total += len(entry)
                    if chunk.doc_name not in sources:
                        sources.append(chunk.doc_name)
                context_parts.append("\n".join(lines))
    except Exception:
        pass

    try:
        from helpers import get_skill_loader
        sl = get_skill_loader()
        if sl is not None:
            skill_ctx, skill_rag_ctx = sl.build_skill_context_with_rag(
                user_msg, max_skill_chars=5000, max_rag_chars=3000, top_k=4
            )
            if skill_ctx:
                context_parts.append("===== 可用技能与函数示例 =====\n" + skill_ctx)
            if skill_rag_ctx:
                context_parts.append("===== 技能绑定知识库 =====\n" + skill_rag_ctx)
    except Exception:
        pass

    if mode == "paper_read":
        system = (
            "你是一位专业的地震学文献解读专家。\n"
            "请基于以下论文内容，用清晰的中文解读、总结或回答用户的问题。\n"
            "回答时请：\n"
            "1. 点明核心方法/创新点\n"
            "2. 解释关键公式或算法（必要时给出 Python 代码示例）\n"
            "3. 说明实验结果与结论\n"
            "4. 指出局限性或未来工作（如有）\n"
        )
    else:
        if context_parts:
            system = (
                "You are SAGE, an expert seismology assistant with deep knowledge of "
                "seismology and data processing.\n"
                "Relevant passages from the knowledge base are provided below. "
                "Use them to answer the question. "
                "If a passage is not directly relevant, rely on your own knowledge instead — "
                "do NOT cite or mention passages that are unrelated to the question.\n"
            )
        else:
            system = (
                "You are SAGE, an expert seismology assistant with deep knowledge of "
                "seismology and data processing.\n"
                "Answer the user's question using your own knowledge. "
                "Be concise and accurate.\n"
            )

    system += _scientific_grounding_policy(bool(data.get("enable_web_search")))
    system = append_user_profile_to_system(system)

    # Think-mode must be model-neutral. OpenAI-compatible reasoning streams are
    # normalized in helpers.llm_stream; non-reasoning local/online models are
    # instructed to emit the same tags so the UI can show/collapse them.
    enable_think = bool(data.get("enable_think", False))
    
    if enable_think:
        system += _think_summary_policy()

    if context_parts:
        system += "\n\n===== Reference passages =====\n" + "\n\n".join(context_parts)

    messages = [{"role": "system", "content": system},
                {"role": "user",   "content": user_msg}]

    history = data.get("history", [])
    if history:
        messages = [messages[0]] + history[-6:] + [messages[-1]]

    return messages, list(set(sources)), llm_cfg


@bp.route('/api/chat/rag/stream', methods=['POST'])
def chat_rag_stream():
    """
    Streaming version of /api/chat/rag.
    Returns text/event-stream SSE with events:
      data: {"type":"sources","sources":[...]}      — sent first
      data: {"type":"chunk","text":"..."}            — one per token
      data: {"type":"done"}
      data: {"type":"error","message":"..."}
    """
    import json as _json

    data = request.json or {}
    if not data.get("message", "").strip():
        return jsonify({"ok": False, "error": "Empty message"}), 400

    images = data.get("images") or []
    messages, sources, llm_cfg = _build_rag_messages(data)

    if not llm_cfg.get("api_base"):
        # Fall back to single-shot response wrapped as SSE
        def _no_backend():
            msg = ("当前没有可用的 LLM 后端。\n"
                   "请在 **LLM 设置** 页面配置后端（Ollama / 在线 API）。")
            yield f"data: {_json.dumps({'type':'sources','sources':[]})}\n\n"
            yield f"data: {_json.dumps({'type':'chunk','text':msg})}\n\n"
            yield f"data: {_json.dumps({'type':'done'})}\n\n"
        return Response(stream_with_context(_no_backend()),
                        mimetype='text/event-stream',
                        headers={'Cache-Control': 'no-cache', 'X-Accel-Buffering': 'no'})

    def generate():
        # Send sources immediately so the UI can show references
        yield f"data: {_json.dumps({'type':'sources','sources':sources})}\n\n"
        try:
            from helpers import llm_stream
            for chunk in llm_stream(messages, llm_cfg, max_tokens=2000,
                                    images=images if images else None):
                yield f"data: {_json.dumps({'type':'chunk','text':chunk})}\n\n"
        except Exception as exc:
            yield f"data: {_json.dumps({'type':'error','message':str(exc)})}\n\n"
        yield f"data: {_json.dumps({'type':'done'})}\n\n"

    return Response(
        stream_with_context(generate()),
        mimetype='text/event-stream',
        headers={'Cache-Control': 'no-cache', 'X-Accel-Buffering': 'no'},
    )


@bp.route('/api/chat/stream', methods=['POST'])
def chat_stream():
    """
    Streaming plain chat (no RAG).  Same SSE format as /api/chat/rag/stream.
    """
    import json as _json

    data    = request.json or {}
    user_msg = data.get("message", "").strip()
    if not user_msg:
        return jsonify({"ok": False, "error": "Empty message"}), 400

    images = data.get("images") or []
    llm_cfg = get_llm_config()

    enable_think = bool(data.get("enable_think", False))
    system = (
        "You are SAGE, an expert seismology assistant with deep knowledge of "
        "seismology, geophysics and data processing.\n"
        "Answer the user's question using your own knowledge. Be concise and accurate.\n"
    )
    system = append_user_profile_to_system(system)
    if enable_think:
        system += _think_summary_policy()

    workspace_path = data.get("workspace", "")
    if workspace_path:
        ws_ctx = inject_workspace_context(user_msg, workspace_path)
        if ws_ctx:
            system += "\n\n===== 本地文件系统 =====\n" + ws_ctx

    project_context = (data.get("project_context") or "").strip()
    if project_context:
        system += "\n\n===== 项目共享上下文 =====\n" + project_context[:4000]

    messages = [{"role": "system", "content": system},
                {"role": "user",   "content": user_msg}]
    history = data.get("history", [])
    if history:
        messages = [messages[0]] + history[-6:] + [messages[-1]]

    def generate():
        yield f"data: {_json.dumps({'type':'sources','sources':[]})}\n\n"
        if not llm_cfg.get("api_base"):
            msg = "当前没有可用的 LLM 后端，请在 LLM 设置页面配置后端。"
            yield f"data: {_json.dumps({'type':'chunk','text':msg})}\n\n"
            yield f"data: {_json.dumps({'type':'done'})}\n\n"
            return
        try:
            from helpers import llm_stream
            for chunk in llm_stream(messages, llm_cfg, max_tokens=2000,
                                    images=images if images else None):
                yield f"data: {_json.dumps({'type':'chunk','text':chunk})}\n\n"
        except Exception as exc:
            yield f"data: {_json.dumps({'type':'error','message':str(exc)})}\n\n"
        yield f"data: {_json.dumps({'type':'done'})}\n\n"

    return Response(
        stream_with_context(generate()),
        mimetype='text/event-stream',
        headers={'Cache-Control': 'no-cache', 'X-Accel-Buffering': 'no'},
    )


# ── 后台异步聊天 Job（切换页面后仍能继续执行并取回结果）────────────────────────

_CHAT_JOB_TTL = 1800   # 30 分钟内可取回结果


def _chat_job_gc():
    cutoff = _time.time() - _CHAT_JOB_TTL
    for k in [k for k, v in list(_chat_jobs.items()) if v.get('ts', 0) < cutoff]:
        _chat_jobs.pop(k, None)


def _build_plain_messages(data: dict):
    """Build messages for plain (no-RAG) chat — mirrors chat_stream() logic."""
    user_msg     = data.get('message', '').strip()
    enable_think = bool(data.get('enable_think', False))
    llm_cfg      = get_llm_config()
    system = (
        'You are SAGE, an expert seismology assistant with deep knowledge of '
        'seismology, geophysics and data processing.\n'
        'Answer the user\'s question using your own knowledge. Be concise and accurate.\n'
    )
    system += _scientific_grounding_policy(bool(data.get("enable_web_search")))
    system = append_user_profile_to_system(system)
    if enable_think:
        system += _think_summary_policy()

    workspace_path = data.get('workspace', '')
    if workspace_path:
        ws_ctx = inject_workspace_context(user_msg, workspace_path)
        if ws_ctx:
            system += '\n\n===== 本地文件系统 =====\n' + ws_ctx

    web_ctx, web_sources = _chat_web_search_context(data, user_msg)
    if web_ctx:
        system += '\n\n===== Web literature/search context =====\n' + web_ctx

    messages = [{'role': 'system', 'content': system},
                {'role': 'user',   'content': user_msg}]
    history = data.get('history', [])
    if history:
        messages = [messages[0]] + history[-6:] + [messages[-1]]
    return messages, web_sources, llm_cfg


@bp.route('/api/chat/submit', methods=['POST'])
def chat_submit():
    """
    Start a non-streaming background chat job.  Returns {job_id} immediately.
    The LLM call runs in a daemon thread — survives page navigation.

    Body fields (same as /api/chat/rag/stream) plus:
      type: 'rag' (default) | 'plain'
    """
    data = request.json or {}
    if not data.get('message', '').strip():
        return jsonify({'ok': False, 'error': 'Empty message'}), 400

    _chat_job_gc()
    job_id = 'chat_' + _uuid.uuid4().hex[:10]
    _chat_jobs[job_id] = {
        'status': 'running', 'answer': '', 'sources': [],
        'error': '', 'cancelled': False, 'ts': _time.time(),
    }

    chat_type = data.get('type', 'rag')
    images    = data.get('images') or []   # list of base64 strings for VL models

    def _run():
        try:
            if chat_type == 'plain':
                messages, sources, llm_cfg = _build_plain_messages(data)
            else:
                messages, sources, llm_cfg = _build_rag_messages(data)

            if not llm_cfg.get('api_base'):
                _chat_jobs[job_id].update(
                    answer='当前没有可用的 LLM 后端。\n请在 **LLM 设置** 页面配置后端（Ollama / 在线 API）。',
                    status='done',
                )
                return

            answer_parts = []
            for chunk in llm_stream(messages, llm_cfg, max_tokens=2000,
                                    images=images if images else None):
                if _chat_jobs[job_id].get('cancelled'):
                    _chat_jobs[job_id].update(status='cancelled', error='已停止当前回复')
                    return
                answer_parts.append(chunk)
                _chat_jobs[job_id].update(
                    answer=''.join(answer_parts),
                    sources=sources,
                    status='running',
                    ts=_time.time(),
                )
            if _chat_jobs[job_id].get('cancelled'):
                _chat_jobs[job_id].update(status='cancelled', error='已停止当前回复')
            else:
                _chat_jobs[job_id].update(answer=''.join(answer_parts), sources=sources, status='done')
        except Exception as exc:
            if _chat_jobs[job_id].get('cancelled'):
                _chat_jobs[job_id].update(status='cancelled', error='已停止当前回复')
            else:
                _chat_jobs[job_id].update(status='error', error=str(exc))

    threading.Thread(target=_run, daemon=True).start()
    return jsonify({'ok': True, 'job_id': job_id})


@bp.route('/api/chat/job/<job_id>', methods=['GET'])
def chat_job_poll(job_id):
    """Poll a background chat job for its result."""
    job = _chat_jobs.get(job_id)
    if not job:
        return jsonify({'ok': False, 'error': 'Job not found or expired'}), 404
    return jsonify({
        'ok':      True,
        'status':  job['status'],   # 'running' | 'done' | 'error'
        'answer':  job['answer'],
        'sources': job['sources'],
        'error':   job['error'],
    })


@bp.route('/api/chat/job/<job_id>/cancel', methods=['POST'])
def chat_job_cancel(job_id):
    job = _chat_jobs.get(job_id)
    if not job:
        return jsonify({'ok': False, 'error': 'Job not found or expired'}), 404
    job['cancelled'] = True
    job['status'] = 'cancelled'
    job['error'] = '已停止当前回复'
    job['ts'] = _time.time()
    return jsonify({'ok': True})


@bp.route('/api/chat/memory', methods=['GET'])
def chat_memory_get():
    return jsonify({
        "ok": True,
        "path": str(USER_PROFILE_MD),
        "content": get_user_profile_context(max_chars=20000),
        "exists": USER_PROFILE_MD.exists(),
        "archive_dir": str(USER_PROFILE_ARCHIVE_DIR),
        "source_path": str(USER_PROFILE_SOURCE_JSON),
    })


@bp.route('/api/chat/memory', methods=['DELETE'])
def chat_memory_delete():
    """Delete persistent user profile / personalization files."""
    deleted = []
    for path in [USER_PROFILE_MD, USER_PROFILE_SOURCE_JSON]:
        try:
            if path.exists():
                path.unlink()
                deleted.append(str(path))
        except Exception as exc:
            return jsonify({"ok": False, "error": str(exc)}), 500
    try:
        if USER_PROFILE_ARCHIVE_DIR.exists():
            for p in USER_PROFILE_ARCHIVE_DIR.glob("user_profile_*.md"):
                p.unlink()
                deleted.append(str(p))
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 500
    return jsonify({"ok": True, "deleted": deleted})


@bp.route('/api/chat/memory/summarize', methods=['POST'])
def chat_memory_summarize():
    """Summarize multiple local conversations into a persistent Markdown user profile."""
    data = request.json or {}
    conversations = data.get("conversations") or []
    if not conversations:
        conversations = _load_persistent_conversations().get("conversations") or []
    if not conversations:
        return jsonify({"ok": False, "error": "No conversations available"}), 400

    snippets = []
    for conv in conversations[:20]:
        title = conv.get("title") or conv.get("id") or "Untitled"
        lines = [f"## Conversation: {title}"]
        for m in (conv.get("history") or [])[-20:]:
            role = m.get("role", "")
            content = str(m.get("content", ""))[:1200]
            if content.strip():
                lines.append(f"- {role}: {content}")
        snippets.append("\n".join(lines))
    corpus = "\n\n".join(snippets)[:30000]

    fallback = (
        "# SAGE User Profile\n\n"
        f"Updated: {datetime.now().isoformat(timespec='seconds')}\n\n"
        "## Inferred identity and work context\n"
        "- The user works on seismology/geophysics workflows and SAGE development.\n\n"
        "## Habits and preferences\n"
        "- Prefers executable workflows, debugging ability, intermediate artifacts, and traceable reasoning.\n"
        "- Often asks for Chinese explanations with practical code support.\n\n"
        "## Knowledge level\n"
        "- Advanced technical user familiar with earthquake monitoring, phase picking, RAG, skills, and coding agents.\n\n"
        "## Intent hints\n"
        "- Explanation-style requests should use QA/RAG.\n"
        "- Requests involving implementation, plotting, detection, processing, or '上述/这个方法 + 代码' should route to coding.\n"
    )

    try:
        llm_cfg = get_llm_config()
        if llm_cfg.get("api_base"):
            prompt = (
                "请根据以下多个对话，生成一份简洁、可长期使用的 Markdown 用户画像。"
                "目标是帮助后续助手判断用户身份、研究方向、习惯、知识水平、偏好的回答深度、"
                "常见意图和路由偏好。不要记录隐私敏感信息，不要逐字复述对话。\n\n"
                "请使用以下结构：\n"
                "# SAGE User Profile\n"
                "## Inferred identity and work context\n"
                "## Research/software interests\n"
                "## Habits and preferences\n"
                "## Knowledge level\n"
                "## Intent and routing hints\n"
                "## Open uncertainties\n\n"
                f"对话材料：\n{corpus}"
            )
            content = llm_call([{"role": "user", "content": prompt}], llm_cfg, max_tokens=1800)
        else:
            content = fallback
    except Exception:
        content = fallback

    _save_user_profile(content, conversations=conversations)
    return jsonify({
        "ok": True,
        "path": str(USER_PROFILE_MD),
        "archive_dir": str(USER_PROFILE_ARCHIVE_DIR),
        "source_path": str(USER_PROFILE_SOURCE_JSON),
        "content": content,
    })
